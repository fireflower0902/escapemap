"""
DB → 전국_방탈출카페_목록.xlsx 재생성 스크립트

실행:
  cd escape-aggregator/backend
  python scripts/export_excel.py
"""

import asyncio
import sys
from pathlib import Path

BACKEND_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(BACKEND_DIR))

from app.database import AsyncSessionLocal
from app.models.cafe import Cafe  # noqa: E402
from sqlalchemy import select

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 없음. 설치 중...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

OUTPUT_PATH = BACKEND_DIR / "data" / "전국_방탈출카페_목록.xlsx"

# 지역 배경색 (ARGB)
REGION_COLORS = {
    "서울": "FFDCE6F1",   # 연파랑
    "경기": "FFD9EAD3",   # 연초록
    "인천": "FFFFF2CC",   # 연노랑
    "부산": "FFFFE6CC",   # 연오렌지
    "대구": "FFFCE5CD",   # 연복숭아
    "광주": "FFEAD1DC",   # 연분홍
    "대전": "FFD0E4FF",   # 연하늘
    "울산": "FFFFE599",   # 연황
    "강원": "FFD5E8D4",   # 연민트
    "경북": "FFFFE6CC",   # 연살구
    "경남": "FFD9D2E9",   # 연보라
    "전북": "FFFFD966",   # 연황금
    "전남": "FFB9E0D4",   # 연청록
    "제주": "FFCFE2F3",   # 하늘
    "충남": "FFEAD1DC",   # 연분홍
    "충북": "FFD5E8D4",   # 연초록
}
DEFAULT_COLOR = "FFF5F5F5"


def get_region(address: str) -> str:
    """주소에서 시도 추출"""
    if not address:
        return "기타"
    parts = address.split()
    return parts[0] if parts else "기타"


def get_region_color(address: str) -> str:
    region = get_region(address)
    for key, color in REGION_COLORS.items():
        if region.startswith(key):
            return color
    return DEFAULT_COLOR


def make_border():
    thin = Side(style="thin", color="FFCCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


async def load_cafes():
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Cafe).order_by(Cafe.address.asc(), Cafe.name.asc())
        )
        return result.scalars().all()


def build_excel(cafes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "전국 방탈출카페"

    # ── 헤더 ──────────────────────────────────────
    headers = ["#", "카페명", "지점명", "주소", "전화번호",
               "공식 홈페이지", "카카오플레이스", "위도", "경도"]
    col_widths = [5, 22, 14, 40, 16, 40, 40, 12, 12]

    header_fill = PatternFill("solid", fgColor="FF2B3A52")
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = make_border()

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── 데이터 행 ──────────────────────────────────
    link_font_real = Font(color="FF0563C1", underline="single", size=10)
    link_font_kakao = Font(color="FF3D7AB5", underline="single", size=10)
    normal_font = Font(size=10)
    left = Alignment(horizontal="left", vertical="center")

    has_homepage = 0
    no_homepage = 0

    for row_idx, cafe in enumerate(cafes, start=2):
        addr = cafe.address or ""
        bg_color = get_region_color(addr)
        row_fill = PatternFill("solid", fgColor=bg_color)

        kakao_url = f"http://place.map.kakao.com/{cafe.id}"

        # 공식 홈페이지: place.map.kakao.com가 아닌 실제 URL
        real_hp = cafe.website_url or ""
        if "place.map.kakao.com" in real_hp:
            real_hp = ""

        if real_hp:
            has_homepage += 1
        else:
            no_homepage += 1

        row_data = [
            row_idx - 1,
            cafe.name or "",
            cafe.branch_name or "",
            addr,
            cafe.phone or "",
            real_hp,        # 공식 홈페이지
            kakao_url,      # 카카오플레이스
            round(cafe.lat, 6) if cafe.lat else "",
            round(cafe.lng, 6) if cafe.lng else "",
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = border

            # URL 열 하이퍼링크
            if col_idx == 6 and real_hp:   # 공식 홈페이지
                cell.hyperlink = real_hp
                cell.font = link_font_real
                cell.alignment = left
            elif col_idx == 7:             # 카카오플레이스
                cell.hyperlink = kakao_url
                cell.font = link_font_kakao
                cell.alignment = left
            elif col_idx in (1, 8, 9):
                cell.font = normal_font
                cell.alignment = center
            else:
                cell.font = normal_font
                cell.alignment = left

    ws.row_dimensions[1].height = 20

    return wb, has_homepage, no_homepage


async def main():
    print("=" * 60)
    print("DB → 전국_방탈출카페_목록.xlsx 재생성")
    print("=" * 60)

    print("\n  DB에서 카페 목록 로딩 중...")
    cafes = await load_cafes()
    print(f"  총 {len(cafes)}개 카페 로딩 완료")

    print("\n  Excel 파일 생성 중...")
    wb, has_hp, no_hp = build_excel(cafes)

    wb.save(OUTPUT_PATH)
    print(f"\n  저장 완료: {OUTPUT_PATH.name}")
    print(f"  공식 홈페이지 보유: {has_hp}개")
    print(f"  공식 홈페이지 없음: {no_hp}개")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
